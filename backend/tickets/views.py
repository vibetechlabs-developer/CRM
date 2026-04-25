from rest_framework.viewsets import ModelViewSet
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from django.db.models import Q
from django.utils import timezone
from rest_framework import filters

from users.models import User
from common.permissions import IsAdminOrAssignedAgent, IsAdminRole, DenyDeleteForManager
from .models import (
    Ticket,
    TicketActivity,
    Note,
    Notification,
    Binder,
    DiscardReopenReminderDismissal,
)
from .serializers import (
    TicketSerializer,
    TicketActivitySerializer,
    NoteSerializer,
    NotificationSerializer,
    BinderSerializer,
)
from .services import (
    auto_assign_ticket,
    discard_reminder_anniversary_for_ticket,
    tickets_in_discard_reopen_reminder_window,
)

# Dashboard preview only; full list via GET /api/tickets/?status=DISCARDED&reopen_reminder=1
DISCARD_REOPEN_REMINDER_PREVIEW_LIMIT = 25


class TicketViewSet(ModelViewSet):

    serializer_class = TicketSerializer
    permission_classes = [IsAuthenticated, DenyDeleteForManager]

    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    search_fields = ['ticket_no', 'client__first_name', 'client__last_name', 'client__email']

    filterset_fields = {
        'status': ['exact'],
        'assigned_to': ['exact'],
        'client': ['exact'],
        'insurance_type': ['exact'],
        'priority': ['exact'],
        'source': ['exact'],
        'created_at': ['date', 'year', 'month'],
    }

    def get_queryset(self):
        qs = Ticket.objects.all().select_related("client", "assigned_to", "policy")
        
        user_role = str(getattr(self.request.user, "role", "") or "").strip().upper()
        is_admin = (
            user_role in {"ADMIN", "MANAGER"}
            or getattr(self.request.user, "is_staff", False)
            or getattr(self.request.user, "is_superuser", False)
        )
        
        if not is_admin:
            qs = qs.filter(assigned_to=self.request.user)

        # ticket_type is special in this app:
        # - UI uses ADJUSTMENT / CUSTOMER_ISSUE, while backend stores adjustment as CHANGES (+ a marker for customer issue).
        ticket_type_param = self.request.query_params.get("ticket_type")
        if ticket_type_param:
            code = ticket_type_param.strip().upper()
            if code == "CUSTOMER_ISSUE":
                qs = qs.filter(ticket_type__in=["CHANGES", "ADJUSTMENT"]).filter(
                    additional_notes__icontains="[Form: Customer Issue]"
                )
            elif code == "ADJUSTMENT":
                qs = qs.filter(ticket_type__in=["CHANGES", "ADJUSTMENT"]).exclude(
                    additional_notes__icontains="[Form: Customer Issue]"
                )
            else:
                qs = qs.filter(ticket_type=code)

        reopen_flag = (self.request.query_params.get("reopen_reminder") or "").strip().lower()
        if reopen_flag in ("1", "true", "yes"):
            viewer = str(getattr(self.request.user, "role", "") or "").strip().upper()
            if viewer in {"ADMIN", "AGENT", "MANAGER"}:
                matches = tickets_in_discard_reopen_reminder_window(
                    qs, timezone.localdate(), exclude_dismissals_for=self.request.user
                )
                qs = qs.filter(id__in=[t.id for t in matches])

        return qs

    def perform_create(self, serializer):
        """
        Create ticket and enforce assignment rules:
        - If AGENT creates: assign to self
        - If ADMIN creates:
            - For RENEWAL/ADJUSTMENT/CANCELLATION: try continuity assignment (same previous agent)
            - Else fallback to auto-assign
        """
        user = self.request.user
        role = getattr(user, "role", None)

        ticket = serializer.save()

        if role == "AGENT":
            if not ticket.assigned_to_id:
                ticket.assigned_to = user
                ticket.save(update_fields=["assigned_to"])
            return

        # ADMIN flow: continuity assignment for non-NEW types
        if ticket.ticket_type in ["RENEWAL", "CHANGES", "ADJUSTMENT", "CANCELLATION"] and not ticket.assigned_to_id:
            previous = (
                Ticket.objects.filter(client=ticket.client, assigned_to__isnull=False)
                .exclude(id=ticket.id)
                .order_by("-created_at")
                .first()
            )
            if previous and previous.assigned_to_id:
                ticket.assigned_to_id = previous.assigned_to_id
                ticket.save(update_fields=["assigned_to"])
            else:
                auto_assign_ticket(ticket)

        # NEW tickets: auto-assign if unassigned (signal also does this, but keep deterministic here)
        if ticket.ticket_type == "NEW" and not ticket.assigned_to_id:
            auto_assign_ticket(ticket)

    @action(detail=False, methods=['get'])
    def stats(self, request):
        from django.db.models import Count, Q, Value
        from django.db.models.functions import Replace, Trim, Upper
        import datetime
        from clients.models import Client

        qs = self.get_queryset()
        completed_statuses = ["COMPLETED", "DONE"]
        discarded_statuses = ["DISCARDED", "DISCARDED_LEADS"]
        
        total_tickets = qs.count()
        completed_tickets = qs.filter(status__in=completed_statuses).count()
        high_priority = qs.filter(priority='HIGH').count()
        active_tickets = qs.exclude(status__in=[*completed_statuses, *discarded_statuses]).count()
        # For AGENT dashboard, only count clients that have tickets assigned to that agent.
        user_role = str(getattr(request.user, "role", "") or "").strip().upper()
        if user_role in {"ADMIN", "MANAGER"}:
            total_clients = Client.objects.count()
        else:
            total_clients = Client.objects.filter(tickets__assigned_to=request.user).distinct().count()

        now = timezone.now()
        def get_month_start(dt, offset):
            m = dt.month - 1 + offset
            y = dt.year + m // 12
            m = m % 12 + 1
            return dt.replace(year=y, month=m, day=1, hour=0, minute=0, second=0, microsecond=0)

        start_this = get_month_start(now, 0)
        start_prev = get_month_start(now, -1)
        start_next = get_month_start(now, 1)

        created_this = qs.filter(created_at__gte=start_this, created_at__lt=start_next)
        created_prev = qs.filter(created_at__gte=start_prev, created_at__lt=start_this)

        tickets_this_month = created_this.count()
        tickets_prev_month = created_prev.count()
        completed_this = created_this.filter(status__in=completed_statuses).count()
        completed_prev = created_prev.filter(status__in=completed_statuses).count()
        high_this = created_this.filter(priority='HIGH').count()
        high_prev = created_prev.filter(priority='HIGH').count()

        def pct_dict(curr, prev):
            if prev <= 0:
                fmt = "0%" if curr <= 0 else "+100%"
                return {"label": fmt, "positive": True}
            val = round(((curr - prev) / prev) * 100)
            return {"label": f"+{val}%" if val >= 0 else f"{val}%", "positive": val >= 0}

        month_stats = {
            "ticketsDelta": pct_dict(tickets_this_month, tickets_prev_month),
            "completedDelta": pct_dict(completed_this, completed_prev),
            "highDelta": pct_dict(high_this, high_prev),
        }

        raw_status_counts = list(qs.values('status').annotate(count=Count('id')))
        status_aliases = {
            "DONE": "COMPLETED",
            "DISCARDED_LEADS": "DISCARDED",
        }
        merged_status_counts = {}
        for row in raw_status_counts:
            status_code = str(row.get("status") or "").upper()
            canonical_status = status_aliases.get(status_code, status_code)
            merged_status_counts[canonical_status] = merged_status_counts.get(canonical_status, 0) + int(row.get("count") or 0)
        status_counts = [{"status": status, "count": count} for status, count in merged_status_counts.items()]
        type_counts = list(qs.values('ticket_type').annotate(count=Count('id')))
        priority_counts = list(qs.values('priority').annotate(count=Count('id')))

        # Be tolerant of legacy/variant values used in older/live data.
        # Normalize ticket_type for bucket math:
        # trim + uppercase + spaces/hyphens => underscores.
        normalized_ticket_type = Upper(
            Replace(
                Replace(Trim("ticket_type"), Value(" "), Value("_")),
                Value("-"),
                Value("_"),
            )
        )
        buckets = []
        for i in range(5, -1, -1):
            d_start = get_month_start(now, -i)
            d_end = get_month_start(now, -i + 1)
            bucket_qs = qs.filter(created_at__gte=d_start, created_at__lt=d_end)
            bucket_qs_normalized = bucket_qs.annotate(ticket_type_norm=normalized_ticket_type)
            buckets.append({
                "key": d_start.strftime("%Y-%m"),
                "month": d_start.strftime("%b"),
                "tickets": bucket_qs.count(),
                "completed": bucket_qs.filter(status__in=completed_statuses).count(),
                # Pattern-based matching is resilient to values like:
                # NEW POLICY, NEW-BUSINESS, RENEWAL REQUEST, POLICY CHANGE, etc.
                "newBusiness": bucket_qs_normalized.filter(ticket_type_norm__startswith="NEW").count(),
                "renewal": bucket_qs_normalized.filter(ticket_type_norm__contains="RENEW").count(),
                "changes": bucket_qs_normalized.filter(
                    Q(ticket_type_norm__contains="CHANGE")
                    | Q(ticket_type_norm__contains="ADJUST")
                    | Q(ticket_type_norm__contains="CUSTOMER_ISSUE")
                    | Q(ticket_type_norm__contains="CUSTOMERISSUE")
                ).count(),
            })

        recent_qs = qs.order_by('-created_at')[:5]
        recent_tickets = self.get_serializer(recent_qs, many=True).data

        payload = {
            "totalTickets": total_tickets,
            "totalClients": total_clients,
            "completedTickets": completed_tickets,
            "highPriority": high_priority,
            "activeTickets": active_tickets,
            "monthStats": month_stats,
            "statusCounts": status_counts,
            "typeCounts": type_counts,
            "priorityCounts": priority_counts,
            "monthlyTrend": buckets,
            "recentTickets": recent_tickets,
        }

        # Year-after discard reminders: ADMIN / AGENT / MANAGER (same broad access as ticket list).
        role = str(getattr(request.user, "role", "") or "").strip().upper()
        if role in {"ADMIN", "AGENT", "MANAGER"}:
            reminder_tickets = tickets_in_discard_reopen_reminder_window(
                qs, timezone.localdate(), exclude_dismissals_for=request.user
            )
            payload["discardReopenReminderCount"] = len(reminder_tickets)
            payload["discardReopenReminders"] = [
                {
                    "id": t.id,
                    "ticket_no": t.ticket_no,
                    "client_name": f"{t.client.first_name or ''} {t.client.last_name or ''}".strip()
                    or (t.client.email or f"Client #{t.client_id}"),
                    "discarded_at": t.discarded_at.isoformat() if t.discarded_at else None,
                }
                for t in reminder_tickets[:DISCARD_REOPEN_REMINDER_PREVIEW_LIMIT]
            ]
        else:
            payload["discardReopenReminderCount"] = 0
            payload["discardReopenReminders"] = []

        return Response(payload)

    @action(detail=False, methods=["post"], url_path="dismiss-discard-reminders")
    def dismiss_discard_reminders(self, request):
        role = str(getattr(request.user, "role", "") or "").strip().upper()
        if role not in {"ADMIN", "AGENT", "MANAGER"}:
            return Response({"detail": "Only agents, managers, and admins can dismiss these reminders."}, status=403)

        qs = self.get_queryset()
        active = tickets_in_discard_reopen_reminder_window(
            qs, timezone.localdate(), exclude_dismissals_for=request.user
        )
        active_ids = {t.id for t in active}

        raw_ids = request.data.get("ticket_ids")
        if raw_ids is not None:
            if not isinstance(raw_ids, list):
                return Response({"detail": "ticket_ids must be a list of integers."}, status=400)
            try:
                want = {int(x) for x in raw_ids}
            except (TypeError, ValueError):
                return Response({"detail": "ticket_ids must be a list of integers."}, status=400)
            to_dismiss_ids = want & active_ids
        else:
            to_dismiss_ids = active_ids

        newly_recorded = 0
        for t in active:
            if t.id not in to_dismiss_ids:
                continue
            anniv = discard_reminder_anniversary_for_ticket(t)
            if anniv is None:
                continue
            _, was_created = DiscardReopenReminderDismissal.objects.get_or_create(
                user=request.user,
                ticket=t,
                reminder_anniversary_on=anniv,
            )
            if was_created:
                newly_recorded += 1

        return Response({"dismissed": len(to_dismiss_ids), "newly_recorded": newly_recorded})

    # Custom API: Change Ticket Status
    @action(detail=True, methods=['post'])
    def change_status(self, request, pk=None):

        ticket = self.get_object()
        new_status = request.data.get('status')
        user_role = str(getattr(request.user, "role", "") or "").strip().upper()
        is_admin = (
            user_role in {"ADMIN", "MANAGER"}
            or getattr(request.user, "is_staff", False)
            or getattr(request.user, "is_superuser", False)
        )

        if not new_status:
            return Response(
                {"error": "Status is required"},
                status=400
            )

        if ticket.status == "COMPLETED" and new_status != "COMPLETED" and not is_admin:
            return Response({"error": "Completed ticket cannot be moved to another stage."}, status=400)
        if ticket.ticket_type == "CANCELLATION" and new_status != "DISCARDED":
            return Response({"error": "Cancellation ticket must stay in Discarded Leads."}, status=400)

        ticket.status = new_status
        ticket._allow_completed_reopen = is_admin
        ticket.save()  # serializer.update will add user-attributed activity when patched via PATCH; here signal won't, so log explicitly
        TicketActivity.objects.create(ticket=ticket, user=request.user, message=f"Status changed to {new_status}.")

        return Response({
            "message": "Ticket status updated successfully",
            "status": ticket.status
        })

    # Custom API: Get Ticket Activities
    @action(detail=True, methods=['get'])
    def activities(self, request, pk=None):

        ticket = self.get_object()

        activities = TicketActivity.objects.filter(
            ticket=ticket
        ).order_by('created_at')

        serializer = TicketActivitySerializer(
            activities,
            many=True
        )

        return Response(serializer.data)

    # Notes
    @action(detail=True, methods=["get", "post"])
    def notes(self, request, pk=None):
        ticket = self.get_object()

        # ADMIN can see all; AGENT only assigned (already enforced by get_queryset)
        if request.method == "GET":
            notes = Note.objects.filter(ticket=ticket).order_by("-created_at")
            return Response(NoteSerializer(notes, many=True).data)

        # POST: create note
        payload = request.data.copy()
        payload["ticket"] = ticket.id
        payload["agent"] = request.user.id
        ser = NoteSerializer(data=payload)
        ser.is_valid(raise_exception=True)
        note = ser.save()
        TicketActivity.objects.create(ticket=ticket, user=request.user, message="Note added.")
        return Response(NoteSerializer(note).data, status=201)

    # Custom API: Auto-assign a specific ticket
    @action(detail=True, methods=['post'])
    def auto_assign(self, request, pk=None):
        """Manually trigger auto-assignment for a specific ticket"""
        ticket = self.get_object()
        
        if ticket.assigned_to:
            return Response({
                "message": "Ticket is already assigned",
                "assigned_to": ticket.assigned_to.username
            })
        
        auto_assign_ticket(ticket)
        ticket.refresh_from_db()
        
        if ticket.assigned_to:
            return Response({
                "message": "Ticket assigned successfully",
                "assigned_to": ticket.assigned_to.username
            })
        else:
            return Response({
                "message": "No available agents found to assign this ticket"
            }, status=404)

    # Custom API: Auto-assign all unassigned tickets
    @action(detail=False, methods=['post'])
    def auto_assign_all(self, request):
        """Auto-assign all unassigned tickets"""
        unassigned_tickets = Ticket.objects.filter(assigned_to__isnull=True)
        total = unassigned_tickets.count()  # capture before loop; queryset re-eval after assignment would return 0
        assigned_count = 0
        failed_count = 0
        
        for ticket in unassigned_tickets:
            old_assigned_to = ticket.assigned_to
            auto_assign_ticket(ticket)
            ticket.refresh_from_db()
            
            if ticket.assigned_to and ticket.assigned_to != old_assigned_to:
                assigned_count += 1
            else:
                failed_count += 1
        
        return Response({
            "message": f"Auto-assignment completed",
            "assigned": assigned_count,
            "failed": failed_count,
            "total_processed": total
        })


class NotificationViewSet(ModelViewSet):
    """
    In-app notifications API (polling-based "real-time").
    """

    serializer_class = NotificationSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.OrderingFilter]
    # Allow filtering notifications by the date they were created.
    # Example: GET /api/notifications/?created_at__date=2026-03-25
    filterset_fields = {"created_at": ["date", "year", "month"]}
    ordering_fields = ["created_at"]

    def get_queryset(self):
        return (
            Notification.objects.filter(user=self.request.user)
            .prefetch_related("ticket", "ticket__activities")
        )

    @action(detail=False, methods=["post"])
    def mark_all_read(self, request):
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return Response({"success": True})

class BinderViewSet(ModelViewSet):
    serializer_class = BinderSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    
    search_fields = ['client_name', 'company_name', 'quote_person', 'binder_person', 'task']
    filterset_fields = []
    ordering_fields = ['binder_date', 'created_at']
    ordering = ['binder_date']  # Default ascending order

    def get_queryset(self):
        qs = Binder.objects.select_related("binder_created_by").all().order_by('binder_date')
        return qs

    def list(self, request, *args, **kwargs):
        response = super().list(request, *args, **kwargs)
        # Prevent intermediary caches from serving user-scoped stale binder lists.
        response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        response["Pragma"] = "no-cache"
        response["Expires"] = "0"
        return response

    def perform_create(self, serializer):
        serializer.save(binder_created_by=self.request.user)

    @action(detail=False, methods=['get'])
    def export(self, request):
        from io import BytesIO
        from django.http import HttpResponse
        from openpyxl import Workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter

        qs = self.filter_queryset(self.get_queryset())

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Binders"

        headers = [
            'Effective Date of Policy', 'Quote Person', 'Binder Person', 'Client Name',
            'Company Name', 'Task', 'Notes', 'Created At'
        ]
        worksheet.append(headers)

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for binder in qs:
            worksheet.append([
                binder.binder_date.strftime("%b %d, %Y") if binder.binder_date else "",
                binder.quote_person,
                binder.binder_person,
                binder.client_name,
                binder.company_name,
                binder.task,
                binder.notes,
                binder.created_at.strftime("%b %d, %Y") if binder.created_at else "",
            ])

        worksheet.freeze_panes = "A2"
        for column in worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                value = "" if cell.value is None else str(cell.value)
                max_length = max(max_length, len(value))
            worksheet.column_dimensions[column_letter].width = min(max(12, max_length + 2), 40)

        output = BytesIO()
        workbook.save(output)
        output.seek(0)

        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="binders_export.xlsx"'
        return response
